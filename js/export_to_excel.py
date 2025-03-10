import openpyxl
import json
import sys
from pathlib import Path

# Путь к файлу Excel
FILE_PATH = "Мои финансы.xlsx"

# Чтение данных из stdin
data = json.loads(sys.stdin.read())

# Загрузка существующей книги или создание новой
if Path(FILE_PATH).exists():
    workbook = openpyxl.load_workbook(FILE_PATH)
else:
    workbook = openpyxl.Workbook()

# Выбор листа "Транзакции" или создание нового
if "Транзакции" in workbook.sheetnames:
    sheet = workbook["Транзакции"]
else:
    sheet = workbook.create_sheet("Транзакции")
    # Добавление заголовков, если лист новый
    sheet.append(["Дата", "Категория", "Счёт", "Тип", "Сумма", "Валюта", "Описание"])

# Добавление новых данных в лист
for transaction in data:
    sheet.append([
        transaction["Дата"],
        transaction["Категория"],
        transaction["Счёт"],
        transaction["Тип"],
        transaction["Сумма"],
        transaction["Валюта"],
        transaction["Описание"]
    ])

# Сохранение книги
workbook.save(FILE_PATH)
print("Данные успешно экспортированы в Excel.")
